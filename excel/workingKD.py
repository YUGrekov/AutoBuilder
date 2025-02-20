import re
import openpyxl as wb
from base_model import BaseModel
from enum import Enum
import traceback

SHIFT = 1
LIST_MODULE = ['CPU', 'PSU', 'CN', 'MN', 'AI', 'AO', 'DI', 'RS', 'DO']


class NameColumn(Enum):
    '''Перечисление статических столбцов таблицы.'''
    ID = 'id'
    TYPE_SIGNAL = 'type_signal'
    TAG = 'tag'
    NAME = 'description'
    SCHEMA = 'schema'
    KLK = 'klk'
    CONTACT = 'contact'
    BASKET = 'basket'
    MODULE = 'module'
    CHANNEL = 'channel'


class DataExel:
    '''Инициализация файла Exel и его таблиц. Выдача массива для записи в базу SQL.'''
    def __init__(self, mainwindow):

        self.m_window = mainwindow
        self.db = self.m_window.tab_1.db_dev.get_database()
        self.logs_msg = self.m_window.logsTextEdit.logs_msg

        self.connect = wb.load_workbook(self.m_window.tab_1.connect.path_to_exel,
                                        read_only=True,
                                        data_only=True)
        BaseModel._meta.database = self.db

    def disconnect_exel(self):
        '''Разрыв связи с Exel.'''
        self.connect.close()

    def read_table(self) -> list:
        '''Список таблиц Exel.'''
        return [sheet.title for sheet in self.connect.worksheets]

    def max_column(self, uso: str):
        """Читаем выбранную таблицу и получаем макс-ное кол-во столбцов."""
        self.sheet = self.connect[uso]
        return self.sheet.max_column

    def read_sheet(self, row: int, column: int) -> str:
        '''Чтение ячейки таблицы Exel.'''
        return self.sheet.cell(row=row, column=column).value

    def read_hat_table(self,
                       uso: str,
                       number_row: int,
                       is_tuple: bool,
                       select_column: tuple = None) -> dict | tuple:
        """Список с названиями столбцов."""
        column = self.max_column(uso)
        hat_tabl = {} if is_tuple else []

        for i in range(int(number_row), int(number_row) + SHIFT):
            for j in range(1, column + SHIFT):
                cell = self.read_sheet(i, j)
                if cell is None:
                    continue

                if is_tuple:
                    for key, value in select_column.items():
                        if value == cell:
                            hat_tabl[key] = j - SHIFT
                else:
                    hat_tabl.append(cell)

        return hat_tabl

    def database_count_row(self):
        '''Вычисляем кол-во строк в таблице Signals в базе SQL.'''
        try:
            count_row = self.m_window.tab_1.db_dev.execute_query('SELECT COUNT(*) FROM signals')[0][0]
        except Exception:
            count_row = 0
        return count_row

    def search_type(self, scheme, type_signal):
        '''Дополнительная проверка на тип сигнала.'''
        return next((value for value in LIST_MODULE if value in str(scheme)), type_signal)
        # for value in LIST_MODULE:
        #     if value in str(scheme):
        #         return value
        # return type_signal

    def sub_str(self, uso, basket, module, channel):
        '''Добавляем теги к резервам.'''
        patterns = {
            r"МНС": "",
            r"ПТ": "",
            r"САР": "",
            r"РП": "",
            r"КЦ": "KC",
            r"с БРУ": "BRU",
            r"БРУ": "BRU",
            r"УСО": "USO"}

        tag = uso
        for pattern, replacement in patterns.items():
            if pattern in uso:
                tag = re.sub(pattern, replacement, tag, flags=re.IGNORECASE)

        # Убираем лишние символы (точки, скобки, пробелы) и добавляем префикс REZ_
        tag = re.sub(r"[\.\s\(\)]", "_", tag)  # Заменяем точки, пробелы, скобки на "_"
        tag = re.sub(r"_+", "_", tag)  # Убираем дублирующиеся "_"
        tag = tag.strip("_")  # Убираем "_" в начале и конце, если есть
        tag = f"REZ_{tag}_"  # Добавляем префикс и суффикс
        return f'{tag}A{basket}_{module}_{channel}'  # Добавляем префикс и суффикс

    def preparation_import(self, uso: str, number_row: int,
                           select_col: tuple) -> list:
        '''Подготовка таблицы к импорту.'''
        data = []

        count_row = self.database_count_row()
        tuple_name = self.read_hat_table(uso, number_row, True, select_col)

        for row in self.sheet.iter_rows(min_row=(int(number_row) + 1)):
            row_data = {col: row[tuple_name[col.value]].value for col in list(NameColumn)[1:]}
            type_s, tag, name, schema, klk, contact, basket, module, channel = row_data.values()

            try:
                # Пропускаем строку, если basket, module или channel отсутствуют
                if not all([basket, module, channel]):
                    continue
                count_row += 1

                if contact is not None:
                    contact = str(contact).replace('.', ',')

                # Генерируем tag, если он отсутствует и schema или type_s не содержат 'RS'
                if tag is None and (schema or type_s) and 'RS' not in (schema, type_s):
                    tag = self.sub_str(uso, basket, module, channel)

                type_s = self.search_type(schema, type_s)

                data.append(dict(id=count_row, type_signal=type_s, uso=uso, tag=tag,
                                 description=name, schema=schema, klk=klk, contact=contact,
                                 basket=basket, module=module, channel=channel))
            except Exception:
                self.m_window.logsTextEdit.logs_msg(f'''Импорт КЗФКП. Таблица: signals, class: DataExel, preparation_import -
                                        Пропуск строки {basket}, {module}, {channel}: {traceback.format_exc()}''', 2)
                continue
        return data


class Import_in_SQL(DataExel):
    '''Запись и обновление сигналов в базе SQL.'''
    def exists_signals(self, row: dict, uso: str):
        '''Проверяем существование сигнала в базе по корзине, модулю, каналу.'''
        from model import Signals
        return Signals.select().where(Signals.uso == uso,
                                      Signals.basket == str(row[NameColumn.BASKET.value]),
                                      Signals.module == str(row[NameColumn.MODULE.value]),
                                      Signals.channel == str(row[NameColumn.CHANNEL.value]))

    def compare_row(self, row_exel: dict, msg: str,
                    object_sql: object, object_exel: str) -> str:
        """Сравнение значений строки из базы SQL и таблицы Exel."""
        if str(object_sql) != str(row_exel[object_exel]):
            return f'{msg}{object_exel}: {row_exel[object_exel]}({object_sql}),'
        return msg

    def record_row(self, row_exel: dict, req_sql: object) -> str:
        """Обновление значения в строке сигнала."""
        from model import Signals

        dop_msg = ''
        for row in req_sql:
            # Список колонок для сравнения
            columns_to_compare = [
                (row.tag, NameColumn.TAG.value),
                (row.description, NameColumn.NAME.value),
                (row.schema, NameColumn.SCHEMA.value),
                (row.klk, NameColumn.KLK.value),
                (row.contact, NameColumn.CONTACT.value)]

            # Сравниваем каждую колонку и обновляем dop_msg
            for value, column in columns_to_compare:
                dop_msg = self.compare_row(row_exel, dop_msg, value, column)

            # Если dop_msg не пустой, обновляем запись в базе данных
            if dop_msg:
                update_data = {
                    NameColumn.TAG.value: row_exel[NameColumn.TAG.value],
                    NameColumn.NAME.value: row_exel[NameColumn.NAME.value],
                    NameColumn.SCHEMA.value: row_exel[NameColumn.SCHEMA.value],
                    NameColumn.KLK.value: row_exel[NameColumn.KLK.value],
                    NameColumn.CONTACT.value: row_exel[NameColumn.CONTACT.value]
                }
                Signals.update(**update_data).where(Signals.id == row.id).execute()

                self.logs_msg(f'''Импорт КД: name = {row_exel[NameColumn.NAME.value]}, id = {row.id},
                              {dop_msg} сигнал обновлен''', 0)

    def work_table(self, clear: bool = False):
        '''Создание или очищение таблицы Signals.'''
        from model import Signals
        try:
            if 'signals' not in self.db.get_tables():
                if not clear:
                    self.db.create_tables([Signals])
                    self.logs_msg('Импорт КД: таблица signals создана', 1)
                else:
                    self.logs_msg('Импорт КД: таблица signals отсутствует', 2)
            else:
                if clear:
                    Signals.delete().execute()
                    self.logs_msg('Импорт КД: таблица signals пустая', 3)
                else:
                    self.logs_msg('Импорт КД: таблица signals уже была создана', 1)
        except Exception:
            self.logs_msg(f'Импорт КД: ошибка импорта {traceback.format_exc()}', 2)

    def database_entry_SQL(self, data: list, uso: str):
        '''Запись новых строк в базу SQL.'''
        from model import Signals
        with self.db.atomic():
            try:
                Signals.insert_many(data).execute()
                self.logs_msg(f'Импорт КД: в таблицу signals добавлено новое УСО {uso}', 1)
            except Exception:
                self.logs_msg(f'Импорт КД: ошибка импорта {traceback.format_exc()}', 2)

    def row_update_SQL(self, data: list, uso: str):
        '''Обновление старой записи, если имеется или запись новой строки.'''
        from model import Signals
        with self.db.atomic():
            try:
                for row_exel in data:
                    exists_s = self.exists_signals(row_exel, uso)
                    if not exists_s:
                        Signals.create(**row_exel)
                        self.logs_msg(f'''Импорт КД: добавлен новый сигнал:
                                      id - {row_exel[NameColumn.ID.value]},
                                      description - {row_exel[NameColumn.NAME.value]},
                                      module - {row_exel[NameColumn.MODULE.value]},
                                      channel - {row_exel[NameColumn.CHANNEL.value]}''', 0)
                    else:
                        self.record_row(row_exel, exists_s)
            except Exception:
                self.logs_msg(f'Импорт КД: ошибка импорта {traceback.format_exc()}', 2)
        self.logs_msg('Импорт КД: обновление завершено', 0)